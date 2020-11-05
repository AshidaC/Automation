import openpyxl


class HomePageData:
    test_Homepage_Data = [
        {"Firstname": "Ashida", "Mail": "ashida.ch@gmail.com", "Password": "ashida@123", "Gender": "Female"},
        {"Firstname": "aslam", "Mail": "aslam@gmail.com", "Password": "aslam@123", "Gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("D:\\PythonFile\\Pythonexceldemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get row
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get coloumn
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
