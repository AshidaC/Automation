import openpyxl

book = openpyxl.load_workbook("D:\\PythonFile\\Pythonexceldemo.xlsx")

sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=1)
print(cell.value)

sheet.cell(row=2, column=2).value = "Ashida"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)

print(sheet.max_column)

print(sheet['A2'].value)

for i in range(1, sheet.max_row + 1):  # to get row
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column + 1):  # to get coloumn
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)